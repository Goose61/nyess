#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Material Takeoff Analyzer for IFC files.
This script analyzes all elements in an IFC file and generates a comprehensive
material takeoff list with dimensions and quantities.
"""

import os
import sys
import logging
import json
import csv
from collections import defaultdict
import ifcopenshell
import ifcopenshell.geom
import ifcopenshell.util.element
import ifcopenshell.util.placement
import ifcopenshell.util.shape
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ifc_database import IFCDatabase

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("material_takeoff.log"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class MaterialTakeoffAnalyzer:
    """
    Analyzes IFC files to generate comprehensive material takeoff lists
    """
    
    def __init__(self, ifc_file_path):
        """
        Initialize the analyzer with an IFC file path.
        
        Args:
            ifc_file_path (str): Path to the IFC file
        """
        self.ifc_file_path = ifc_file_path
        self.logger = logging.getLogger(__name__)
        self.logged_material_ids = set()  # Track which material IDs we've already logged errors for
        
        try:
            self.ifc_file = ifcopenshell.open(ifc_file_path)
            self.logger.info(f"Successfully loaded IFC file: {ifc_file_path}")
            self.logger.info(f"IFC schema: {self.ifc_file.schema}")
            
            # Initialize database
            self.db = IFCDatabase()
            self.ifc_file_id = self.db.store_ifc_file(ifc_file_path, self.ifc_file.schema)
        except Exception as e:
            self.logger.error(f"Failed to load IFC file: {e}")
            raise
        
        # Initialize settings for geometry processing
        self.settings = ifcopenshell.geom.settings()
        self.settings.set(self.settings.USE_WORLD_COORDS, True)
        
        # Initialize material takeoff data structure
        self.results = {
            'element_types': defaultdict(lambda: {
                'count': 0,
                'total_volume': 0.0,
                'total_area': 0.0,
                'materials': defaultdict(lambda: {
                    'count': 0,
                    'volume': 0.0,
                    'area': 0.0,
                    'properties': defaultdict(str),
                    'grades': [],
                    'specifications': [],
                    'material_type': '',
                    'category': '',
                    'description': '',
                    'dimensions': []
                }),
                'dimensions': []
            }),
            'materials': defaultdict(lambda: {
                'count': 0,
                'total_volume': 0.0,
                'total_area': 0.0,
                'properties': defaultdict(str),
                'grades': [],
                'specifications': [],
                'material_type': '',
                'category': '',
                'description': '',
                'element_types': [],
                'dimensions': []
            })
        }
    
    def calculate_volume_and_area(self, shape):
        """Calculate volume and area from shape geometry."""
        try:
            if not shape or not shape.geometry:
                return 0.0, 0.0
                
            # Get vertices and faces
            verts = shape.geometry.verts
            faces = shape.geometry.faces
            
            if not verts or not faces:
                return 0.0, 0.0
                
            # Convert to numpy arrays
            verts_array = np.array(verts).reshape(-1, 3)
            faces_array = np.array(faces).reshape(-1, 3)
            
            # Calculate area using triangulated faces
            area = 0.0
            for face in faces_array:
                v1 = verts_array[face[0]]
                v2 = verts_array[face[1]]
                v3 = verts_array[face[2]]
                
                # Calculate face normal
                edge1 = v2 - v1
                edge2 = v3 - v1
                normal = np.cross(edge1, edge2)
                face_area = np.linalg.norm(normal) / 2.0
                area += face_area
            
            # Estimate volume using bounding box
            bbox = self.calculate_bounding_box(verts)
            if bbox:
                dims = bbox['bounding_box']['dimensions']
                volume = dims[0] * dims[1] * dims[2]
            else:
                volume = 0.0
            
            return volume, area
            
        except Exception as e:
            self.logger.warning(f"Error calculating volume and area: {str(e)}")
            return 0.0, 0.0

    def analyze_all_elements(self):
        """Analyze all elements in the IFC file."""
        total_elements = len(self.ifc_file.by_type('IfcProduct'))
        processed_elements = 0
        batch_size = 100  # Process elements in batches
        
        self.logger.info(f"Analyzing {total_elements} elements")
        
        # Create a structure to track unique elements by dimensions and material
        element_catalog = defaultdict(lambda: {
            'count': 0,
            'volume': 0.0,
            'area': 0.0,
            'elements': [],
            'dimensions': None,
            'material_data': None
        })
        
        try:
            for product in self.ifc_file.by_type('IfcProduct'):
                try:
                    processed_elements += 1
                    if processed_elements % batch_size == 0:
                        self.logger.info(f"Processed {processed_elements}/{total_elements} elements ({(processed_elements/total_elements)*100:.1f}%)")
                    
                    # Skip non-physical elements
                    if not product.is_a('IfcElement'):
                        continue
                    
                    element_type = product.is_a()
                    self.results['element_types'][element_type]['count'] += 1
                    
                    # Get materials
                    materials = self.get_materials_with_properties(product)
                    
                    # Try to get geometry
                    try:
                        shape = ifcopenshell.geom.create_shape(self.settings, product)
                        if shape:
                            # Calculate volume and area using our custom method
                            volume, area = self.calculate_volume_and_area(shape)
                            
                            # Calculate bounding box
                            bbox = self.calculate_bounding_box(shape.geometry.verts)
                            if not bbox:
                                continue
                                
                            # Normalize dimensions (sort them by size)
                            dimensions = sorted(bbox['bounding_box']['dimensions'])
                            length, width, height = dimensions[2], dimensions[1], dimensions[0]
                            
                            # Round dimensions to nearest millimeter (3 decimal places in meters)
                            length = round(length, 3)
                            width = round(width, 3)
                            height = round(height, 3)
                            
                            # Update element type totals
                            self.results['element_types'][element_type]['total_volume'] += volume
                            self.results['element_types'][element_type]['total_area'] += area
                            self.results['element_types'][element_type]['dimensions'].append(bbox)
                            
                            # For each material, add this element to the catalog
                            for material_name, material_data in materials.items():
                                # Generate a unique key for this element type + dimension + material
                                dim_key = f"{element_type}|{material_name}|{length}x{width}x{height}"
                                
                                # Add to catalog of unique elements
                                element_catalog[dim_key]['count'] += 1
                                element_catalog[dim_key]['volume'] += volume
                                element_catalog[dim_key]['area'] += area
                                
                                if not element_catalog[dim_key]['dimensions']:
                                    element_catalog[dim_key]['dimensions'] = {
                                        'length': length,
                                        'width': width,
                                        'height': height
                                    }
                                
                                if not element_catalog[dim_key]['material_data']:
                                    element_catalog[dim_key]['material_data'] = material_data
                                
                                element_catalog[dim_key]['elements'].append({
                                    'id': product.id(),
                                    'name': product.Name if hasattr(product, 'Name') else '',
                                    'volume': volume,
                                    'area': area,
                                    'length': length,
                                    'width': width,
                                    'height': height
                                })
                                
                                # Update element type material data
                                self.results['element_types'][element_type]['materials'][material_name]['count'] += 1
                                self.results['element_types'][element_type]['materials'][material_name]['volume'] += volume
                                self.results['element_types'][element_type]['materials'][material_name]['area'] += area
                                self.results['element_types'][element_type]['materials'][material_name]['properties'].update(material_data['properties'])
                                self.results['element_types'][element_type]['materials'][material_name]['grades'].extend(material_data['grades'])
                                self.results['element_types'][element_type]['materials'][material_name]['specifications'].extend(material_data['specifications'])
                                self.results['element_types'][element_type]['materials'][material_name]['material_type'] = material_data['material_type']
                                self.results['element_types'][element_type]['materials'][material_name]['category'] = material_data['category']
                                self.results['element_types'][element_type]['materials'][material_name]['description'] = material_data['description']
                                if bbox:
                                    self.results['element_types'][element_type]['materials'][material_name]['dimensions'].append(bbox)
                                
                                # Update global material data
                                self.results['materials'][material_name]['count'] += 1
                                self.results['materials'][material_name]['total_volume'] += volume
                                self.results['materials'][material_name]['total_area'] += area
                                self.results['materials'][material_name]['properties'].update(material_data['properties'])
                                self.results['materials'][material_name]['grades'].extend(material_data['grades'])
                                self.results['materials'][material_name]['specifications'].extend(material_data['specifications'])
                                self.results['materials'][material_name]['material_type'] = material_data['material_type']
                                self.results['materials'][material_name]['category'] = material_data['category']
                                self.results['materials'][material_name]['description'] = material_data['description']
                                self.results['materials'][material_name]['element_types'].append(element_type)
                                if bbox:
                                    self.results['materials'][material_name]['dimensions'].append(bbox)
                    except Exception as e:
                        self.logger.warning(f"Error processing geometry for element {product.id()}: {str(e)}")
                        continue
                    
                except Exception as e:
                    self.logger.warning(f"Error processing element {product.id()}: {str(e)}")
                    continue
                
        except KeyboardInterrupt:
            self.logger.warning("Analysis interrupted by user. Saving partial results...")
        
        # Store the element catalog in the results
        self.results['element_catalog'] = dict(element_catalog)
        
        # Calculate summary statistics
        self.calculate_summary_statistics(self.results)
        
        return self.results
    
    def get_materials_with_properties(self, element):
        """Extract material information with properties from an element."""
        materials = defaultdict(lambda: {
            'properties': defaultdict(str),
            'grades': [],
            'specifications': [],
            'material_type': '',
            'category': '',
            'description': ''
        })
        
        try:
            # Use ifcopenshell's utility function to get material
            material_data = ifcopenshell.util.element.get_material(element)
            if not material_data:
                return materials
                
            # Handle different material types
            if isinstance(material_data, (list, tuple)):
                # Handle material lists
                for mat in material_data:
                    if isinstance(mat, str):
                        # Handle string materials
                        materials[mat]['material_type'] = 'Material'
                        materials[mat]['description'] = 'String material'
                    else:
                        self._process_material(mat, materials)
            elif isinstance(material_data, str):
                # Handle string material
                materials[material_data]['material_type'] = 'Material'
                materials[material_data]['description'] = 'String material'
            elif hasattr(material_data, 'is_a'):
                # Handle single material
                self._process_material(material_data, materials)
                
        except Exception as e:
            self.logger.warning(f"Error getting materials for element {element.id()}: {str(e)}")
        
        return materials
        
    def _process_material(self, material, materials_dict):
        """Process a single material and extract its properties."""
        try:
            # Check if material is a string, which happens in some IFC files
            if isinstance(material, str):
                mat_name = material
                materials_dict[mat_name]['material_type'] = 'Material'
                materials_dict[mat_name]['description'] = 'String material'
                return
                
            if material.is_a('IfcMaterial'):
                mat_name = material.Name
                materials_dict[mat_name]['material_type'] = 'Material'
                materials_dict[mat_name]['category'] = material.Category if hasattr(material, 'Category') else ''
                materials_dict[mat_name]['description'] = material.Description if hasattr(material, 'Description') else ''
                
                # Get material properties using ifcopenshell utility
                props = ifcopenshell.util.element.get_properties(material)
                for pset_name, pset_props in props.items():
                    for prop_name, prop_value in pset_props.items():
                        # Check for grade and specification properties
                        if 'grade' in prop_name.lower():
                            if str(prop_value) not in materials_dict[mat_name]['grades']:
                                materials_dict[mat_name]['grades'].append(str(prop_value))
                        elif 'specification' in prop_name.lower():
                            if str(prop_value) not in materials_dict[mat_name]['specifications']:
                                materials_dict[mat_name]['specifications'].append(str(prop_value))
                        else:
                            materials_dict[mat_name]['properties'][prop_name] = str(prop_value)
                            
            elif material.is_a('IfcMaterialLayerSetUsage'):
                for layer in material.ForLayerSet.MaterialLayers:
                    if layer.Material and not isinstance(layer.Material, str):
                        mat_name = layer.Material.Name
                        materials_dict[mat_name]['material_type'] = 'Material Layer'
                        materials_dict[mat_name]['category'] = layer.Material.Category if hasattr(layer.Material, 'Category') else ''
                        materials_dict[mat_name]['description'] = layer.Material.Description if hasattr(layer.Material, 'Description') else ''
                        materials_dict[mat_name]['properties']['LayerThickness'] = str(layer.LayerThickness)
                        
                        # Get material properties
                        props = ifcopenshell.util.element.get_properties(layer.Material)
                        for pset_name, pset_props in props.items():
                            for prop_name, prop_value in pset_props.items():
                                if 'grade' in prop_name.lower():
                                    if str(prop_value) not in materials_dict[mat_name]['grades']:
                                        materials_dict[mat_name]['grades'].append(str(prop_value))
                                elif 'specification' in prop_name.lower():
                                    if str(prop_value) not in materials_dict[mat_name]['specifications']:
                                        materials_dict[mat_name]['specifications'].append(str(prop_value))
                                else:
                                    materials_dict[mat_name]['properties'][prop_name] = str(prop_value)
                                    
            elif material.is_a('IfcMaterialProfileSetUsage'):
                for profile in material.ForProfileSet.MaterialProfiles:
                    if profile.Material and not isinstance(profile.Material, str):
                        mat_name = profile.Material.Name
                        materials_dict[mat_name]['material_type'] = 'Material Profile'
                        materials_dict[mat_name]['category'] = profile.Material.Category if hasattr(profile.Material, 'Category') else ''
                        materials_dict[mat_name]['description'] = profile.Material.Description if hasattr(profile.Material, 'Description') else ''
                        
                        # Get material properties
                        props = ifcopenshell.util.element.get_properties(profile.Material)
                        for pset_name, pset_props in props.items():
                            for prop_name, prop_value in pset_props.items():
                                if 'grade' in prop_name.lower():
                                    if str(prop_value) not in materials_dict[mat_name]['grades']:
                                        materials_dict[mat_name]['grades'].append(str(prop_value))
                                elif 'specification' in prop_name.lower():
                                    if str(prop_value) not in materials_dict[mat_name]['specifications']:
                                        materials_dict[mat_name]['specifications'].append(str(prop_value))
                                else:
                                    materials_dict[mat_name]['properties'][prop_name] = str(prop_value)
                                    
        except Exception as e:
            material_id = material.id() if hasattr(material, 'id') else 'unknown'
            if material_id not in self.logged_material_ids:
                self.logger.warning(f"Error processing material {material_id}: {str(e)}")
                self.logged_material_ids.add(material_id)
    
    def calculate_bounding_box(self, vertices):
        """
        Calculate bounding box from vertex data.
        
        Args:
            vertices (list): List of vertex coordinates
            
        Returns:
            dict: Bounding box information
        """
        try:
            if not vertices:
                return None
            
            # Convert to numpy array
            verts_array = np.array(vertices).reshape(-1, 3)
            
            # Calculate min and max coordinates
            min_coords = verts_array.min(axis=0)
            max_coords = verts_array.max(axis=0)
            
            # Calculate dimensions
            dimensions = max_coords - min_coords
            
            return {
                'bounding_box': {
                    'min': min_coords.tolist(),
                    'max': max_coords.tolist(),
                    'dimensions': dimensions.tolist()
                }
            }
        except Exception as e:
            self.logger.warning(f"Error calculating bounding box: {e}")
            return None
    
    def calculate_summary_statistics(self, results):
        """
        Calculate summary statistics for each element type.
        """
        for element_type, data in results['element_types'].items():
            if data['count'] > 0:
                # Calculate average dimensions from bounding boxes
                if data['dimensions']:
                    # Extract dimensions from all bounding boxes
                    lengths = []
                    widths = []
                    heights = []
                    
                    for bbox in data['dimensions']:
                        if bbox and 'bounding_box' in bbox:
                            dims = bbox['bounding_box']['dimensions']
                            lengths.append(dims[0])
                            widths.append(dims[1])
                            heights.append(dims[2])
                    
                    # Calculate averages
                    if lengths:
                        data['avg_length'] = sum(lengths) / len(lengths)
                    if widths:
                        data['avg_width'] = sum(widths) / len(widths)
                    if heights:
                        data['avg_height'] = sum(heights) / len(heights)
                
                # Calculate total quantities
                data['total_quantity'] = {
                    'volume': data['total_volume'],
                    'area': data['total_area'],
                    'count': data['count']
                }
                
                # Calculate material statistics
                for material_name, material_data in data['materials'].items():
                    if material_data['dimensions']:
                        # Extract dimensions from all bounding boxes
                        lengths = []
                        widths = []
                        heights = []
                        
                        for bbox in material_data['dimensions']:
                            if bbox and 'bounding_box' in bbox:
                                dims = bbox['bounding_box']['dimensions']
                                lengths.append(dims[0])
                                widths.append(dims[1])
                                heights.append(dims[2])
                        
                        # Calculate averages
                        if lengths:
                            material_data['avg_length'] = sum(lengths) / len(lengths)
                        if widths:
                            material_data['avg_width'] = sum(widths) / len(widths)
                        if heights:
                            material_data['avg_height'] = sum(heights) / len(heights)
    
    def save_results(self, output_format='all'):
        """
        Save material takeoff results to file.
        
        Args:
            output_format (str): Output format (json, csv, excel, or all)
        """
        base_filename = os.path.splitext(os.path.basename(self.ifc_file_path))[0]
        
        if output_format in ['json', 'all']:
            try:
                output_path = f"{base_filename}_material_takeoff.json"
                with open(output_path, 'w') as f:
                    json.dump(self.results, f, indent=4)
                self.logger.info(f"Material takeoff saved to {output_path}")
            except Exception as e:
                self.logger.error(f"Error saving JSON file: {str(e)}")
        
        if output_format in ['excel', 'all']:
            # Create Excel workbook
            excel_path = f"{base_filename}_material_takeoff.xlsx"
            wb = Workbook()
            
            # Create Material Takeoff sheet
            takeoff_sheet = wb.active
            takeoff_sheet.title = "Detailed Material Takeoff"
            
            # Add headers for Material Takeoff (professional format)
            headers = [
                'Element Type', 'Material Name', 'Grade', 'Specification', 
                'Length (m)', 'Width (m)', 'Height (m)', 'Quantity', 
                'Unit', 'Volume Each (m³)', 'Total Volume (m³)', 
                'Weight Each (kg)', 'Total Weight (kg)', 'Comments'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = takeoff_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Add data to Material Takeoff sheet
            row = 2
            current_element_type = None
            
            # First sort the catalog by element type, then by material
            sorted_keys = sorted(self.results['element_catalog'].keys(), 
                             key=lambda x: (x.split('|')[0], x.split('|')[1]))
            
            for key in sorted_keys:
                element_type, material_name, _ = key.split('|', 2)
                data = self.results['element_catalog'][key]
                material_data = data['material_data']
                dimensions = data['dimensions']
                
                # Create element type headers (groups)
                if current_element_type != element_type:
                    current_element_type = element_type
                    takeoff_sheet.cell(row=row, column=1, value=element_type)
                    takeoff_sheet.cell(row=row, column=1).font = Font(bold=True)
                    for col in range(1, len(headers) + 1):
                        takeoff_sheet.cell(row=row, column=col).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                    row += 1
                
                # Calculate density for weight calculation (default to 7850 kg/m³ for steel if not specified)
                density = 7850  # Default density (steel)
                for prop_name, prop_value in material_data['properties'].items():
                    if 'density' in prop_name.lower() and prop_value:
                        try:
                            density = float(prop_value)
                            break
                        except ValueError:
                            pass
                
                # Calculate weight
                volume_each = data['volume'] / data['count'] if data['count'] > 0 else 0
                weight_each = volume_each * density
                total_weight = data['volume'] * density
                
                # Prepare comments (can include material type, category, etc.)
                comments = material_data['description']
                
                # Write data
                takeoff_sheet.cell(row=row, column=1, value=element_type)
                takeoff_sheet.cell(row=row, column=2, value=material_name)
                takeoff_sheet.cell(row=row, column=3, value=', '.join(material_data['grades']))
                takeoff_sheet.cell(row=row, column=4, value=', '.join(material_data['specifications']))
                takeoff_sheet.cell(row=row, column=5, value=dimensions['length'])
                takeoff_sheet.cell(row=row, column=6, value=dimensions['width'])
                takeoff_sheet.cell(row=row, column=7, value=dimensions['height'])
                takeoff_sheet.cell(row=row, column=8, value=data['count'])
                takeoff_sheet.cell(row=row, column=9, value="ea")
                takeoff_sheet.cell(row=row, column=10, value=round(volume_each, 3))
                takeoff_sheet.cell(row=row, column=11, value=round(data['volume'], 3))
                takeoff_sheet.cell(row=row, column=12, value=round(weight_each, 1))
                takeoff_sheet.cell(row=row, column=13, value=round(total_weight, 1))
                takeoff_sheet.cell(row=row, column=14, value=comments)
                
                # Add border to all cells in the row
                for col in range(1, len(headers) + 1):
                    takeoff_sheet.cell(row=row, column=col).border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'),
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                
                row += 1
            
            # Create Material Summary sheet
            summary_sheet = wb.create_sheet("Material Summary")
            
            # Add headers for Material Summary
            headers = ['Material Name', 'Total Count', 'Total Volume (m³)', 'Total Area (m²)', 'Total Weight (kg)',
                      'Grade', 'Specification', 'Material Type', 'Category', 'Description']
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Add data to Material Summary sheet
            row = 2
            for material_name, data in self.results['materials'].items():
                # Calculate density for weight calculation (default to 7850 kg/m³ for steel if not specified)
                density = 7850  # Default density (steel)
                for prop_name, prop_value in data['properties'].items():
                    if 'density' in prop_name.lower() and prop_value:
                        try:
                            density = float(prop_value)
                            break
                        except ValueError:
                            pass
                
                # Calculate total weight
                total_weight = data['total_volume'] * density
                
                # Write data to separate cells
                summary_sheet.cell(row=row, column=1, value=material_name)
                summary_sheet.cell(row=row, column=2, value=data['count'])
                summary_sheet.cell(row=row, column=3, value=round(data['total_volume'], 3))
                summary_sheet.cell(row=row, column=4, value=round(data['total_area'], 3))
                summary_sheet.cell(row=row, column=5, value=round(total_weight, 1))
                summary_sheet.cell(row=row, column=6, value=', '.join(sorted(data['grades'])))
                summary_sheet.cell(row=row, column=7, value=', '.join(sorted(data['specifications'])))
                summary_sheet.cell(row=row, column=8, value=data['material_type'])
                summary_sheet.cell(row=row, column=9, value=data['category'])
                summary_sheet.cell(row=row, column=10, value=data['description'])
                
                # Add border to all cells in the row
                for col in range(1, len(headers) + 1):
                    summary_sheet.cell(row=row, column=col).border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'),
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                
                row += 1
            
            # Create Element Type Summary sheet
            element_sheet = wb.create_sheet("Element Type Summary")
            
            # Add headers for Element Type Summary
            headers = ['Element Type', 'Count', 'Total Volume (m³)', 'Total Area (m²)', 'Total Weight (kg)',
                      'Materials']
            for col, header in enumerate(headers, 1):
                cell = element_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Add data to Element Type Summary sheet
            row = 2
            for element_type, data in self.results['element_types'].items():
                if data['count'] > 0:
                    # Calculate total weight (using default density of steel)
                    total_weight = data['total_volume'] * 7850  # Default to steel density
                    
                    # Write data to separate cells
                    element_sheet.cell(row=row, column=1, value=element_type)
                    element_sheet.cell(row=row, column=2, value=data['count'])
                    element_sheet.cell(row=row, column=3, value=round(data['total_volume'], 3))
                    element_sheet.cell(row=row, column=4, value=round(data['total_area'], 3))
                    element_sheet.cell(row=row, column=5, value=round(total_weight, 1))
                    element_sheet.cell(row=row, column=6, value=', '.join(data['materials'].keys()))
                    
                    # Add border to all cells in the row
                    for col in range(1, len(headers) + 1):
                        element_sheet.cell(row=row, column=col).border = Border(
                            left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'), 
                            bottom=Side(style='thin')
                        )
                    
                    row += 1
            
            # Create Element Details sheet
            details_sheet = wb.create_sheet("Element Details")
            
            # Add headers for Element Details
            headers = ['Element Type', 'Material Name', 'Element ID', 'Length (m)', 'Width (m)', 'Height (m)',
                      'Volume (m³)', 'Area (m²)', 'Weight (kg)']
            for col, header in enumerate(headers, 1):
                cell = details_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Add individual element data to Element Details sheet
            row = 2
            for key in sorted_keys:
                element_type, material_name, _ = key.split('|', 2)
                data = self.results['element_catalog'][key]
                
                for element in data['elements']:
                    # Calculate density (default to steel)
                    density = 7850  # Default density (steel)
                    for prop_name, prop_value in data['material_data']['properties'].items():
                        if 'density' in prop_name.lower() and prop_value:
                            try:
                                density = float(prop_value)
                                break
                            except ValueError:
                                pass
                    
                    # Calculate weight
                    weight = element['volume'] * density
                    
                    # Write data
                    details_sheet.cell(row=row, column=1, value=element_type)
                    details_sheet.cell(row=row, column=2, value=material_name)
                    details_sheet.cell(row=row, column=3, value=element['id'])
                    details_sheet.cell(row=row, column=4, value=element['length'])
                    details_sheet.cell(row=row, column=5, value=element['width'])
                    details_sheet.cell(row=row, column=6, value=element['height'])
                    details_sheet.cell(row=row, column=7, value=round(element['volume'], 3))
                    details_sheet.cell(row=row, column=8, value=round(element['area'], 3))
                    details_sheet.cell(row=row, column=9, value=round(weight, 1))
                    
                    # Add border to all cells in the row
                    for col in range(1, len(headers) + 1):
                        details_sheet.cell(row=row, column=col).border = Border(
                            left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'), 
                            bottom=Side(style='thin')
                        )
                    
                    row += 1
            
            # Adjust column widths for all sheets
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save workbook with error handling and retry mechanism
            max_attempts = 3
            attempt = 0
            saved = False
            
            while attempt < max_attempts and not saved:
                try:
                    if attempt > 0:
                        # Try with a different filename if first attempt failed
                        excel_path = f"{base_filename}_material_takeoff_{attempt}.xlsx"
                        
                    wb.save(excel_path)
                    self.logger.info(f"Material takeoff Excel file saved to {excel_path}")
                    saved = True
                except PermissionError:
                    attempt += 1
                    self.logger.warning(f"Permission denied when saving Excel file. Attempt {attempt}/{max_attempts}.")
                    if attempt == max_attempts:
                        self.logger.error(f"Failed to save Excel file after {max_attempts} attempts. The file may be open in another application.")
                except Exception as e:
                    self.logger.error(f"Error saving Excel file: {str(e)}")
                    break
        
        if output_format in ['csv', 'all']:
            try:
                # Create summary CSV
                summary_path = f"{base_filename}_material_takeoff_summary.csv"
                with open(summary_path, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Element Type', 'Count', 'Total Volume (m³)', 
                                   'Total Area (m²)', 'Materials'])
                    
                    for element_type, data in self.results['element_types'].items():
                        writer.writerow([
                            element_type,
                            data['count'],
                            f"{data['total_volume']:.2f}",
                            f"{data['total_area']:.2f}",
                            ', '.join(data['materials'].keys())
                        ])
                
                self.logger.info(f"Material takeoff summary saved to {summary_path}")
                
                # Create detailed CSV
                details_path = f"{base_filename}_material_takeoff_details.csv"
                with open(details_path, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Element Type', 'ID', 'Name', 'Volume (m³)', 
                                   'Area (m²)', 'Length (m)', 'Width (m)', 'Height (m)',
                                   'Materials', 'Material Properties'])
                    
                    for element_type, data in self.results['element_types'].items():
                        for element in data.get('elements', []):
                            writer.writerow([
                                element_type,
                                element.get('id'),
                                element.get('name'),
                                f"{element.get('volume', 0):.2f}",
                                f"{element.get('area', 0):.2f}",
                                f"{element.get('length', 0):.2f}",
                                f"{element.get('width', 0):.2f}",
                                f"{element.get('height', 0):.2f}",
                                ', '.join(element.get('materials', {}).keys()),
                                json.dumps(element.get('properties', {}))
                            ])
                
                self.logger.info(f"Material takeoff details saved to {details_path}")
            except Exception as e:
                self.logger.error(f"Error saving CSV file: {str(e)}")
        
        # Close database connection
        self.db.close()

def main():
    """Main function to run the material takeoff analyzer."""
    if len(sys.argv) != 2:
        logger.error("Please provide an IFC file path")
        logger.info("Usage: python material_takeoff.py path/to/your_model.ifc")
        return 1
    
    ifc_file_path = sys.argv[1]
    if not os.path.isfile(ifc_file_path):
        logger.error(f"IFC file not found: {ifc_file_path}")
        return 1
    
    try:
        # Create analyzer and process elements
        analyzer = MaterialTakeoffAnalyzer(ifc_file_path)
        results = analyzer.analyze_all_elements()
        
        # Save results in all formats
        analyzer.save_results('all')
        
        # Display summary
        logger.info("\nMaterial Takeoff Summary:")
        for element_type, data in results['element_types'].items():
            if data['count'] > 0:
                logger.info(f"\n{element_type}:")
                logger.info(f"  Count: {data['count']}")
                logger.info(f"  Total Volume: {data['total_volume']:.2f} m³")
                logger.info(f"  Total Area: {data['total_area']:.2f} m²")
                logger.info("  Materials:")
                for material_name, material_data in data['materials'].items():
                    logger.info(f"    - {material_name}: {material_data['count']} elements")
                    if material_data['grades']:
                        logger.info(f"      Grades: {', '.join(material_data['grades'])}")
        
        return 0
    
    except Exception as e:
        logger.exception(f"Error analyzing IFC file: {e}")
        return 1

if __name__ == "__main__":
    sys.exit(main()) 