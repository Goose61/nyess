#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Material Takeoff Analyzer for IFC files.
This script analyzes all elements in an IFC file and generates a comprehensive
material takeoff list with dimensions and quantities.
"""

import os
import sys
import logging
import json
import csv
from collections import defaultdict
import ifcopenshell
import ifcopenshell.geom
import ifcopenshell.util.element
import ifcopenshell.util.placement
import ifcopenshell.util.shape
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ifc_database import IFCDatabase
import logging.handlers
import tempfile
import time

# Create logs directory if it doesn't exist
os.makedirs("logs", exist_ok=True)

# Configure logging with a rotating file handler
log_file = os.path.join("logs", "material_takeoff.log")
handler = logging.handlers.RotatingFileHandler(
    log_file,
    maxBytes=10*1024*1024,  # 10MB
    backupCount=5,
    encoding='utf-8'
)
handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))

# Configure root logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.addHandler(handler)
logger.addHandler(logging.StreamHandler(sys.stdout))

# Prevent duplicate logging
logger.propagate = False

class MaterialTakeoffAnalyzer:
    """
    Analyzes IFC files to generate comprehensive material takeoff lists
    """
    
    def __init__(self, ifc_file_path):
        """
        Initialize the analyzer with an IFC file path.
        
        Args:
            ifc_file_path (str): Path to the IFC file
        """
        self.ifc_file_path = ifc_file_path
        self.logger = logger
        self.logged_material_ids = set()  # Track which material IDs we've already logged errors for
        
        try:
            self.ifc_file = ifcopenshell.open(ifc_file_path)
            self.logger.info(f"Successfully loaded IFC file: {ifc_file_path}")
            self.logger.info(f"IFC schema: {self.ifc_file.schema}")
            
            # Initialize database
            self.db = IFCDatabase()
            self.ifc_file_id = self.db.store_ifc_file(ifc_file_path, self.ifc_file.schema)
        except Exception as e:
            self.logger.error(f"Failed to load IFC file: {e}")
            raise
        
        # Initialize settings for geometry processing
        self.settings = ifcopenshell.geom.settings()
        self.settings.set(self.settings.USE_WORLD_COORDS, True)
        
        # Initialize material takeoff data structure
        self.results = {
            'element_types': defaultdict(lambda: {
                'count': 0,
                'total_volume': 0.0,
                'total_area': 0.0,
                'materials': defaultdict(lambda: {
                    'count': 0,
                    'volume': 0.0,
                    'area': 0.0,
                    'properties': defaultdict(str),
                    'grades': [],
                    'specifications': [],
                    'material_type': '',
                    'category': '',
                    'description': '',
                    'dimensions': []
                }),
                'dimensions': []
            }),
            'materials': defaultdict(lambda: {
                'count': 0,
                'total_volume': 0.0,
                'total_area': 0.0,
                'properties': defaultdict(str),
                'grades': [],
                'specifications': [],
                'material_type': '',
                'category': '',
                'description': '',
                'element_types': [],
                'dimensions': []
            })
        }
    
    def calculate_volume_and_area(self, shape):
        """Calculate volume and area from shape geometry."""
        try:
            if not shape or not shape.geometry:
                return 0.0, 0.0
                
            # Get vertices and faces
            verts = shape.geometry.verts
            faces = shape.geometry.faces
            
            if not verts or not faces:
                return 0.0, 0.0
                
            # Convert to numpy arrays
            verts_array = np.array(verts).reshape(-1, 3)
            faces_array = np.array(faces).reshape(-1, 3)
            
            # Calculate area using triangulated faces
            area = 0.0
            for face in faces_array:
                v1 = verts_array[face[0]]
                v2 = verts_array[face[1]]
                v3 = verts_array[face[2]]
                
                # Calculate face normal
                edge1 = v2 - v1
                edge2 = v3 - v1
                normal = np.cross(edge1, edge2)
                face_area = np.linalg.norm(normal) / 2.0
                area += face_area
            
            # Estimate volume using bounding box
            bbox = self.calculate_bounding_box(verts)
            if bbox:
                dims = bbox['bounding_box']['dimensions']
                volume = dims[0] * dims[1] * dims[2]
            else:
                volume = 0.0
            
            return volume, area
            
        except Exception as e:
            self.logger.warning(f"Error calculating volume and area: {str(e)}")
            return 0.0, 0.0

    def analyze_all_elements(self):
        """Analyze all elements in the IFC file."""
        total_elements = len(self.ifc_file.by_type('IfcProduct'))
        processed_elements = 0
        batch_size = 100  # Process elements in batches
        
        self.logger.info(f"Analyzing {total_elements} elements")
        
        # Create a structure to track unique elements by dimensions and material
        # Using a regular dict instead of defaultdict for better type checking
        element_catalog = {}
        
        try:
            for product in self.ifc_file.by_type('IfcProduct'):
                try:
                    processed_elements += 1
                    if processed_elements % batch_size == 0:
                        self.logger.info(f"Processed {processed_elements}/{total_elements} elements ({(processed_elements/total_elements)*100:.1f}%)")
                    
                    # Skip non-physical elements
                    if not product.is_a('IfcElement'):
                        continue
                    
                    element_type = product.is_a()
                    self.results['element_types'][element_type]['count'] += 1
                    
                    # Get materials
                    materials = self.get_materials_with_properties(product)
                    
                    # Try to get geometry
                    try:
                        shape = ifcopenshell.geom.create_shape(self.settings, product)
                        if shape:
                            # Calculate volume and area using our custom method
                            volume, area = self.calculate_volume_and_area(shape)
                            
                            # Calculate bounding box
                            bbox = self.calculate_bounding_box(shape.geometry.verts)
                            if not bbox:
                                continue
                                
                            # Normalize dimensions (sort them by size)
                            dimensions = sorted(bbox['bounding_box']['dimensions'])
                            length, width, height = dimensions[2], dimensions[1], dimensions[0]
                            
                            # Round dimensions to nearest millimeter (3 decimal places in meters)
                            length = round(length, 3)
                            width = round(width, 3)
                            height = round(height, 3)
                            
                            # Update element type totals
                            self.results['element_types'][element_type]['total_volume'] += volume
                            self.results['element_types'][element_type]['total_area'] += area
                            self.results['element_types'][element_type]['dimensions'].append(bbox)
                            
                            # For each material, add this element to the catalog
                            for material_name, material_data in materials.items():
                                # Generate a unique key for this element type + dimension + material
                                dim_key = f"{element_type}|{material_name}|{length}x{width}x{height}"
                                
                                # Initialize the catalog entry if it doesn't exist yet
                                if dim_key not in element_catalog:
                                    element_catalog[dim_key] = {
                                        'count': 0,
                                        'volume': 0.0,
                                        'area': 0.0,
                                        'elements': [],
                                        'dimensions': None,
                                        'material_data': None
                                    }
                                
                                # Add to catalog of unique elements
                                element_catalog[dim_key]['count'] += 1
                                element_catalog[dim_key]['volume'] += volume
                                element_catalog[dim_key]['area'] += area
                                
                                if element_catalog[dim_key]['dimensions'] is None:
                                    element_catalog[dim_key]['dimensions'] = {
                                        'length': length,
                                        'width': width,
                                        'height': height
                                    }
                                
                                if element_catalog[dim_key]['material_data'] is None:
                                    element_catalog[dim_key]['material_data'] = material_data
                                
                                # Add element to the list
                                element_catalog[dim_key]['elements'].append({
                                    'id': product.id(),
                                    'name': product.Name if hasattr(product, 'Name') else '',
                                    'volume': volume,
                                    'area': area,
                                    'length': length,
                                    'width': width,
                                    'height': height
                                })
                                
                                # Update element type material data
                                if material_name not in self.results['element_types'][element_type]['materials']:
                                    self.results['element_types'][element_type]['materials'][material_name] = {
                                        'count': 0,
                                        'volume': 0.0,
                                        'area': 0.0,
                                        'properties': defaultdict(str),
                                        'grades': [],
                                        'specifications': [],
                                        'material_type': '',
                                        'category': '',
                                        'description': '',
                                        'dimensions': []
                                    }
                                
                                # Now perform the updates with proper initialization
                                self.results['element_types'][element_type]['materials'][material_name]['count'] += 1
                                self.results['element_types'][element_type]['materials'][material_name]['volume'] += volume
                                self.results['element_types'][element_type]['materials'][material_name]['area'] += area
                                self.results['element_types'][element_type]['materials'][material_name]['properties'].update(material_data['properties'])
                                
                                # Safe extension of lists
                                if isinstance(material_data.get('grades'), list):
                                    self.results['element_types'][element_type]['materials'][material_name]['grades'].extend(material_data['grades'])
                                if isinstance(material_data.get('specifications'), list):
                                    self.results['element_types'][element_type]['materials'][material_name]['specifications'].extend(material_data['specifications'])
                                    
                                self.results['element_types'][element_type]['materials'][material_name]['material_type'] = material_data['material_type']
                                self.results['element_types'][element_type]['materials'][material_name]['category'] = material_data['category']
                                self.results['element_types'][element_type]['materials'][material_name]['description'] = material_data['description']
                                if bbox:
                                    self.results['element_types'][element_type]['materials'][material_name]['dimensions'].append(bbox)
                                
                                # Update global material data
                                # Initialize first if it doesn't exist
                                if material_name not in self.results['materials']:
                                    self.results['materials'][material_name] = {
                                        'count': 0,
                                        'total_volume': 0.0,
                                        'total_area': 0.0,
                                        'properties': defaultdict(str),
                                        'grades': [],
                                        'specifications': [],
                                        'material_type': '',
                                        'category': '',
                                        'description': '',
                                        'element_types': [],
                                        'dimensions': []
                                    }
                                    
                                self.results['materials'][material_name]['count'] += 1
                                self.results['materials'][material_name]['total_volume'] += volume
                                self.results['materials'][material_name]['total_area'] += area
                                self.results['materials'][material_name]['properties'].update(material_data['properties'])
                                
                                # Safely extend lists
                                if isinstance(material_data.get('grades'), list):
                                    self.results['materials'][material_name]['grades'].extend(material_data['grades'])
                                if isinstance(material_data.get('specifications'), list):
                                    self.results['materials'][material_name]['specifications'].extend(material_data['specifications'])
                                    
                                self.results['materials'][material_name]['material_type'] = material_data['material_type']
                                self.results['materials'][material_name]['category'] = material_data['category']
                                self.results['materials'][material_name]['description'] = material_data['description']
                                
                                # Safely append to element_types list
                                if isinstance(self.results['materials'][material_name]['element_types'], list):
                                    self.results['materials'][material_name]['element_types'].append(element_type)
                                
                                # Safely append to dimensions list
                                if bbox and isinstance(self.results['materials'][material_name]['dimensions'], list):
                                    self.results['materials'][material_name]['dimensions'].append(bbox)
                    except Exception as e:
                        self.logger.warning(f"Error processing geometry for element {product.id()}: {str(e)}")
                        continue
                    
                except Exception as e:
                    self.logger.warning(f"Error processing element {product.id()}: {str(e)}")
                    continue
                
        except KeyboardInterrupt:
            self.logger.warning("Analysis interrupted by user. Saving partial results...")
        
        # Store the element catalog in the results
        self.results['element_catalog'] = dict(element_catalog)
        
        # Calculate summary statistics and store the updated results
        self.results = self.calculate_summary_statistics(self.results)
        
        return self.results
    
    def get_materials_with_properties(self, element):
        """Extract material information with properties from an element."""
        materials = defaultdict(lambda: {
            'properties': defaultdict(str),
            'grades': [],
            'specifications': [],
            'material_type': '',
            'category': '',
            'description': ''
        })
        
        try:
            # Use ifcopenshell's utility function to get material, with better error handling
            try:
                material_data = ifcopenshell.util.element.get_material(element)
            except AttributeError as ae:
                # Handle possible attribute errors from ifcopenshell
                self.logger.warning(f"AttributeError getting material for element {element.id()}: {str(ae)}")
                # Return empty materials rather than crashing
                return materials
            
            if not material_data:
                return materials
                
            # Handle different material types
            if isinstance(material_data, (list, tuple)):
                # Handle material lists
                for mat in material_data:
                    if isinstance(mat, str):
                        # Handle string materials
                        materials[mat]['material_type'] = 'String Material'
                        materials[mat]['description'] = 'String material from IFC'
                        # Set properties directly for string materials
                        materials[mat]['properties'] = defaultdict(str, {
                            'Name': mat,
                            'Type': 'String Material'
                        })
                    else:
                        try:
                            self._process_material(mat, materials)
                        except AttributeError as ae:
                            self.logger.warning(f"AttributeError processing material in element {element.id()}: {str(ae)}")
                            # Skip this material and continue
                            continue
            elif isinstance(material_data, str):
                # Handle string material
                materials[material_data]['material_type'] = 'String Material'
                materials[material_data]['description'] = 'String material from IFC'
                # Set properties directly for string materials
                materials[material_data]['properties'] = defaultdict(str, {
                    'Name': material_data,
                    'Type': 'String Material'
                })
            elif hasattr(material_data, 'is_a'):
                # Handle single material
                try:
                    self._process_material(material_data, materials)
                except AttributeError as ae:
                    self.logger.warning(f"AttributeError processing material in element {element.id()}: {str(ae)}")
                    # Try with a fallback approach
                    if hasattr(material_data, 'Name'):
                        mat_name = material_data.Name
                        materials[mat_name]['material_type'] = 'Material'
                        materials[mat_name]['description'] = 'Fallback material'
                        # Set properties directly for fallback materials
                        materials[mat_name]['properties'] = defaultdict(str, {
                            'Name': mat_name,
                            'Type': 'Material'
                        })
            
        except Exception as e:
            element_id = element.id() if hasattr(element, 'id') else 'unknown'
            self.logger.warning(f"Error getting materials for element {element_id}: {str(e)}")
        
        return materials
        
    def _process_material(self, material, materials_dict):
        """Process a single material and extract its properties."""
        try:
            # Handle string materials first
            if isinstance(material, str):
                mat_name = material
                if mat_name not in materials_dict:
                    materials_dict[mat_name] = {
                        'material_type': 'String Material',
                        'category': '',
                        'description': 'String material from IFC',
                        'properties': defaultdict(str, {
                            'Name': mat_name,
                            'Type': 'String Material'
                        }),
                        'grades': [],
                        'specifications': [],
                        'count': 0,
                        'volume': 0.0,
                        'area': 0.0,
                        'dimensions': []
                    }
                return
            
            # For IFC material objects, ensure they have required attributes
            if not hasattr(material, 'is_a'):
                self.logger.warning(f"Material object has no is_a method: {material}")
                return
                
            # Get material name safely
            if not hasattr(material, 'Name'):
                mat_name = f"Unknown-{id(material)}"
                self.logger.warning(f"Material has no Name attribute, using {mat_name}")
            else:
                mat_name = material.Name if material.Name else f"Unnamed-{id(material)}"
            
            # Initialize material entry if not exists
            if mat_name not in materials_dict:
                materials_dict[mat_name] = {
                    'material_type': '',
                    'category': '',
                    'description': '',
                    'properties': defaultdict(str),
                    'grades': [],
                    'specifications': [],
                    'count': 0,
                    'volume': 0.0,
                    'area': 0.0,
                    'dimensions': []
                }
            
            # Process by type with better error handling
            if material.is_a('IfcMaterial'):
                materials_dict[mat_name]['material_type'] = 'Material'
                materials_dict[mat_name]['category'] = material.Category if hasattr(material, 'Category') else ''
                materials_dict[mat_name]['description'] = material.Description if hasattr(material, 'Description') else ''
                
                # Get material properties using ifcopenshell utility
                try:
                    props = ifcopenshell.util.element.get_properties(material)
                    for pset_name, pset_props in props.items():
                        for prop_name, prop_value in pset_props.items():
                            # Check for grade and specification properties
                            if 'grade' in prop_name.lower():
                                if str(prop_value) not in materials_dict[mat_name]['grades']:
                                    materials_dict[mat_name]['grades'].append(str(prop_value))
                            elif 'specification' in prop_name.lower():
                                if str(prop_value) not in materials_dict[mat_name]['specifications']:
                                    materials_dict[mat_name]['specifications'].append(str(prop_value))
                            else:
                                materials_dict[mat_name]['properties'][prop_name] = str(prop_value)
                except Exception as e:
                    # Only log the error once per material
                    if mat_name not in self.logged_material_ids:
                        self.logger.warning(f"Error getting properties for material {mat_name}: {str(e)}")
                        self.logged_material_ids.add(mat_name)
                    # Add basic properties even if we couldn't get the full set
                    materials_dict[mat_name]['properties'].update({
                        'Name': mat_name,
                        'Type': 'IfcMaterial'
                    })
                        
            elif material.is_a('IfcMaterialLayerSetUsage'):
                try:
                    if not hasattr(material, 'ForLayerSet') or not material.ForLayerSet:
                        self.logger.warning(f"IfcMaterialLayerSetUsage has no ForLayerSet: {mat_name}")
                        materials_dict[mat_name]['material_type'] = 'Material Layer Set'
                        materials_dict[mat_name]['description'] = 'Missing layer set'
                        return
                    
                    if not hasattr(material.ForLayerSet, 'MaterialLayers'):
                        self.logger.warning(f"IfcMaterialLayerSet has no MaterialLayers: {mat_name}")
                        materials_dict[mat_name]['material_type'] = 'Material Layer Set'
                        materials_dict[mat_name]['description'] = 'Missing material layers'
                        return
                    
                    for layer in material.ForLayerSet.MaterialLayers:
                        if not layer or not hasattr(layer, 'Material') or not layer.Material:
                            continue
                        
                        if isinstance(layer.Material, str):
                            layer_mat_name = layer.Material
                            if layer_mat_name not in materials_dict:
                                materials_dict[layer_mat_name] = {
                                    'material_type': 'String Material Layer',
                                    'category': '',
                                    'description': 'String material layer',
                                    'properties': defaultdict(str, {
                                        'Name': layer_mat_name,
                                        'Type': 'String Material Layer'
                                    }),
                                    'grades': [],
                                    'specifications': [],
                                    'count': 0,
                                    'volume': 0.0,
                                    'area': 0.0,
                                    'dimensions': []
                                }
                        else:
                            layer_mat_name = layer.Material.Name if hasattr(layer.Material, 'Name') and layer.Material.Name else f"Layer-{id(layer)}"
                            materials_dict[layer_mat_name]['material_type'] = 'Material Layer'
                            materials_dict[layer_mat_name]['category'] = layer.Material.Category if hasattr(layer.Material, 'Category') else ''
                            materials_dict[layer_mat_name]['description'] = layer.Material.Description if hasattr(layer.Material, 'Description') else ''
                            
                            if hasattr(layer, 'LayerThickness'):
                                materials_dict[layer_mat_name]['properties']['LayerThickness'] = str(layer.LayerThickness)
                            
                            # Get material properties
                            try:
                                props = ifcopenshell.util.element.get_properties(layer.Material)
                                for pset_name, pset_props in props.items():
                                    for prop_name, prop_value in pset_props.items():
                                        if 'grade' in prop_name.lower():
                                            if str(prop_value) not in materials_dict[layer_mat_name]['grades']:
                                                materials_dict[layer_mat_name]['grades'].append(str(prop_value))
                                        elif 'specification' in prop_name.lower():
                                            if str(prop_value) not in materials_dict[layer_mat_name]['specifications']:
                                                materials_dict[layer_mat_name]['specifications'].append(str(prop_value))
                                        else:
                                            materials_dict[layer_mat_name]['properties'][prop_name] = str(prop_value)
                            except Exception as e:
                                # Only log the error once per material
                                if layer_mat_name not in self.logged_material_ids:
                                    self.logger.warning(f"Error getting properties for layer material {layer_mat_name}: {str(e)}")
                                    self.logged_material_ids.add(layer_mat_name)
                                # Add basic properties even if we couldn't get the full set
                                materials_dict[layer_mat_name]['properties'].update({
                                    'Name': layer_mat_name,
                                    'Type': 'Material Layer'
                                })
                except Exception as e:
                    # Only log the error once per material
                    if mat_name not in self.logged_material_ids:
                        self.logger.warning(f"Error processing IfcMaterialLayerSetUsage {mat_name}: {str(e)}")
                        self.logged_material_ids.add(mat_name)
                                
            elif material.is_a('IfcMaterialProfileSetUsage'):
                try:
                    if not hasattr(material, 'ForProfileSet') or not material.ForProfileSet:
                        self.logger.warning(f"IfcMaterialProfileSetUsage has no ForProfileSet: {mat_name}")
                        materials_dict[mat_name]['material_type'] = 'Material Profile Set'
                        materials_dict[mat_name]['description'] = 'Missing profile set'
                        return
                    
                    if not hasattr(material.ForProfileSet, 'MaterialProfiles'):
                        self.logger.warning(f"IfcMaterialProfileSet has no MaterialProfiles: {mat_name}")
                        materials_dict[mat_name]['material_type'] = 'Material Profile Set'
                        materials_dict[mat_name]['description'] = 'Missing material profiles'
                        return
                    
                    for profile in material.ForProfileSet.MaterialProfiles:
                        if not profile or not hasattr(profile, 'Material') or not profile.Material:
                            continue
                        
                        if isinstance(profile.Material, str):
                            profile_mat_name = profile.Material
                            if profile_mat_name not in materials_dict:
                                materials_dict[profile_mat_name] = {
                                    'material_type': 'String Material Profile',
                                    'category': '',
                                    'description': 'String material profile',
                                    'properties': defaultdict(str, {
                                        'Name': profile_mat_name,
                                        'Type': 'String Material Profile'
                                    }),
                                    'grades': [],
                                    'specifications': [],
                                    'count': 0,
                                    'volume': 0.0,
                                    'area': 0.0,
                                    'dimensions': []
                                }
                        else:
                            profile_mat_name = profile.Material.Name if hasattr(profile.Material, 'Name') and profile.Material.Name else f"Profile-{id(profile)}"
                            materials_dict[profile_mat_name]['material_type'] = 'Material Profile'
                            materials_dict[profile_mat_name]['category'] = profile.Material.Category if hasattr(profile.Material, 'Category') else ''
                            materials_dict[profile_mat_name]['description'] = profile.Material.Description if hasattr(profile.Material, 'Description') else ''
                            
                            # Get material properties
                            try:
                                props = ifcopenshell.util.element.get_properties(profile.Material)
                                for pset_name, pset_props in props.items():
                                    for prop_name, prop_value in pset_props.items():
                                        if 'grade' in prop_name.lower():
                                            if str(prop_value) not in materials_dict[profile_mat_name]['grades']:
                                                materials_dict[profile_mat_name]['grades'].append(str(prop_value))
                                        elif 'specification' in prop_name.lower():
                                            if str(prop_value) not in materials_dict[profile_mat_name]['specifications']:
                                                materials_dict[profile_mat_name]['specifications'].append(str(prop_value))
                                        else:
                                            materials_dict[profile_mat_name]['properties'][prop_name] = str(prop_value)
                            except Exception as e:
                                # Only log the error once per material
                                if profile_mat_name not in self.logged_material_ids:
                                    self.logger.warning(f"Error getting properties for profile material {profile_mat_name}: {str(e)}")
                                    self.logged_material_ids.add(profile_mat_name)
                                # Add basic properties even if we couldn't get the full set
                                materials_dict[profile_mat_name]['properties'].update({
                                    'Name': profile_mat_name,
                                    'Type': 'Material Profile'
                                })
                except Exception as e:
                    # Only log the error once per material
                    if mat_name not in self.logged_material_ids:
                        self.logger.warning(f"Error processing IfcMaterialProfileSetUsage {mat_name}: {str(e)}")
                        self.logged_material_ids.add(mat_name)
                                    
        except Exception as e:
            material_id = getattr(material, 'id', lambda: 'unknown')()
            if material_id not in self.logged_material_ids:
                self.logger.warning(f"Error processing material {material_id}: {str(e)}")
                self.logged_material_ids.add(material_id)
    
    def calculate_bounding_box(self, vertices):
        """
        Calculate bounding box from vertex data.
        
        Args:
            vertices (list): List of vertex coordinates
            
        Returns:
            dict: Bounding box information
        """
        try:
            if not vertices:
                return None
            
            # Convert to numpy array
            verts_array = np.array(vertices).reshape(-1, 3)
            
            # Calculate min and max coordinates
            min_coords = verts_array.min(axis=0)
            max_coords = verts_array.max(axis=0)
            
            # Calculate dimensions
            dimensions = max_coords - min_coords
            
            return {
                'bounding_box': {
                    'min': min_coords.tolist(),
                    'max': max_coords.tolist(),
                    'dimensions': dimensions.tolist()
                }
            }
        except Exception as e:
            self.logger.warning(f"Error calculating bounding box: {e}")
            return None
    
    def calculate_summary_statistics(self, results):
        """
        Calculate summary statistics for each element type.
        
        Args:
            results (dict): The results dictionary to update with statistics
            
        Returns:
            dict: The updated results dictionary with statistics
        """
        # Safety check for None or invalid results
        if not isinstance(results, dict) or 'element_types' not in results:
            self.logger.warning("Invalid results dictionary passed to calculate_summary_statistics")
            return results
        
        for element_type, data in results['element_types'].items():
            if not isinstance(data, dict) or data.get('count', 0) <= 0:
                continue
            
            # Calculate average dimensions from bounding boxes
            if isinstance(data.get('dimensions'), list) and data['dimensions']:
                # Extract dimensions from all bounding boxes
                lengths = []
                widths = []
                heights = []
                
                for bbox in data['dimensions']:
                    if isinstance(bbox, dict) and 'bounding_box' in bbox:
                        dims = bbox['bounding_box'].get('dimensions', [0, 0, 0])
                        if len(dims) >= 3:  # Make sure we have 3 dimensions
                            lengths.append(dims[0])
                            widths.append(dims[1])
                            heights.append(dims[2])
                
                # Calculate averages
                if lengths:
                    data['avg_length'] = sum(lengths) / len(lengths)
                if widths:
                    data['avg_width'] = sum(widths) / len(widths)
                if heights:
                    data['avg_height'] = sum(heights) / len(heights)
            
            # Calculate total quantities
            data['total_quantity'] = {
                'volume': data.get('total_volume', 0),
                'area': data.get('total_area', 0),
                'count': data.get('count', 0)
            }
            
            # Calculate material statistics
            if isinstance(data.get('materials'), dict):
                for material_name, material_data in data['materials'].items():
                    if not isinstance(material_data, dict):
                        continue
                        
                    if isinstance(material_data.get('dimensions'), list) and material_data['dimensions']:
                        # Extract dimensions from all bounding boxes
                        lengths = []
                        widths = []
                        heights = []
                        
                        for bbox in material_data['dimensions']:
                            if isinstance(bbox, dict) and 'bounding_box' in bbox:
                                dims = bbox['bounding_box'].get('dimensions', [0, 0, 0])
                                if len(dims) >= 3:  # Make sure we have 3 dimensions
                                    lengths.append(dims[0])
                                    widths.append(dims[1])
                                    heights.append(dims[2])
                        
                        # Calculate averages
                        if lengths:
                            material_data['avg_length'] = sum(lengths) / len(lengths)
                        if widths:
                            material_data['avg_width'] = sum(widths) / len(widths)
                        if heights:
                            material_data['avg_height'] = sum(heights) / len(heights)
        
        # Return the modified results
        return results
    
    def save_results(self, output_format='all', results_folder=None):
        """
        Save material takeoff results to file.
        
        Args:
            output_format (str): Output format (json, csv, excel, or all)
            results_folder (str): Optional folder to save results, defaults to same folder as IFC file
        """
        # Get base filename from the IFC path
        base_filename = os.path.splitext(os.path.basename(self.ifc_file_path))[0]
        
        # Determine output directory - prefer results_folder if provided
        if results_folder:
            output_dir = results_folder
        else:
            output_dir = os.path.dirname(self.ifc_file_path)
        
        self.logger.info(f"Saving results to directory: {output_dir}")
        self.logger.info(f"Base filename: {base_filename}")
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Test if directory is writable
        try:
            test_file = os.path.join(output_dir, '.test_write')
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            self.logger.info(f"Directory {output_dir} is writable")
        except Exception as e:
            self.logger.error(f"Directory {output_dir} is not writable: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            
            # Try to use temp directory as fallback
            import tempfile
            output_dir = tempfile.gettempdir()
            self.logger.info(f"Using temp directory instead: {output_dir}")
        
        # Dictionary to track saved files
        saved_files = {}
        
        if output_format in ['json', 'all']:
            try:
                json_file = f"{base_filename}_material_takeoff.json"
                output_path = os.path.join(output_dir, json_file)
                self.logger.info(f"Saving JSON to: {output_path}")
                
                with open(output_path, 'w') as f:
                    json.dump(self.results, f, indent=4)
                
                # Verify file was created
                if os.path.exists(output_path):
                    self.logger.info(f"Material takeoff saved to {output_path}")
                    saved_files['json_file'] = json_file
                else:
                    self.logger.error(f"Failed to create JSON file at {output_path}")
            except Exception as e:
                self.logger.error(f"Error saving JSON file: {str(e)}")
                import traceback
                self.logger.error(traceback.format_exc())
        
        if output_format in ['excel', 'all']:
            # Create Excel workbook with absolute path
            excel_file = f"{base_filename}_material_takeoff.xlsx"
            excel_path = os.path.join(output_dir, excel_file)
            self.logger.info(f"Saving Excel to: {excel_path}")
            
            if self.save_to_excel(excel_path):
                saved_files['excel_file'] = excel_file
            
            # Verify file was created
            if os.path.exists(excel_path):
                self.logger.info(f"Excel file verified at {excel_path}")
            else:
                self.logger.error(f"Failed to create Excel file at {excel_path}")
        
        if output_format in ['csv', 'all']:
            try:
                # Create summary CSV with absolute path
                summary_file = f"{base_filename}_material_takeoff_summary.csv"
                details_file = f"{base_filename}_material_takeoff_details.csv"
                summary_path = os.path.join(output_dir, summary_file)
                details_path = os.path.join(output_dir, details_file)
                self.logger.info(f"Saving summary CSV to: {summary_path}")
                
                with open(summary_path, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Element Type', 'Count', 'Total Volume (m³)', 
                                   'Total Area (m²)', 'Materials'])
                    
                    for element_type, data in self.results['element_types'].items():
                        writer.writerow([
                            element_type,
                            data['count'],
                            f"{data['total_volume']:.2f}",
                            f"{data['total_area']:.2f}",
                            ', '.join(data['materials'].keys())
                        ])
                
                # Verify file was created
                if os.path.exists(summary_path):
                    self.logger.info(f"Material takeoff summary saved to {summary_path}")
                    saved_files['summary_file'] = summary_file
                else:
                    self.logger.error(f"Failed to create summary CSV at {summary_path}")
                
                self.logger.info(f"Saving details CSV to: {details_path}")
                with open(details_path, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Element Type', 'ID', 'Name', 'Volume (m³)', 
                                   'Area (m²)', 'Length (m)', 'Width (m)', 'Height (m)',
                                   'Materials', 'Material Properties'])
                    
                    for element_type, data in self.results['element_types'].items():
                        for element in data.get('elements', []):
                            writer.writerow([
                                element_type,
                                element.get('id'),
                                element.get('name'),
                                f"{element.get('volume', 0):.2f}",
                                f"{element.get('area', 0):.2f}",
                                f"{element.get('length', 0):.2f}",
                                f"{element.get('width', 0):.2f}",
                                f"{element.get('height', 0):.2f}",
                                ', '.join(element.get('materials', {}).keys()),
                                json.dumps(element.get('properties', {}))
                            ])
                
                # Verify file was created
                if os.path.exists(details_path):
                    self.logger.info(f"Material takeoff details saved to {details_path}")
                    saved_files['details_file'] = details_file
                else:
                    self.logger.error(f"Failed to create details CSV at {details_path}")
            except Exception as e:
                self.logger.error(f"Error saving CSV file: {str(e)}")
                import traceback
                self.logger.error(traceback.format_exc())
        
        # Save the output directory path
        saved_files['output_dir'] = output_dir
        
        # Close database connection
        self.db.close()
        
        # Return paths of generated files for reference
        return saved_files

    def save_to_excel(self, output_file):
        """Save results to an Excel file."""
        try:
            self.logger.info(f"Creating Excel workbook for {output_file}")
            wb = Workbook()
            
            # Create element sheet
            self._create_element_sheet(wb)
            
            # Create material sheet
            self._create_material_sheet(wb)
            
            # Create summary sheet
            self._create_summary_sheet(wb)
            
            # Create a directory if it doesn't exist
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                self.logger.info(f"Creating directory: {output_dir}")
                os.makedirs(output_dir, exist_ok=True)
            
            # Check if directory is writable
            try:
                test_file = os.path.join(output_dir, '.test_write')
                with open(test_file, 'w') as f:
                    f.write('test')
                os.remove(test_file)
                self.logger.info(f"Directory {output_dir} is writable")
            except Exception as e:
                self.logger.error(f"Directory {output_dir} is not writable: {str(e)}")
                
                # Try to use a temp directory instead
                import tempfile
                temp_dir = tempfile.gettempdir()
                output_file = os.path.join(temp_dir, os.path.basename(output_file))
                self.logger.info(f"Using temp directory instead: {output_file}")
            
            # Save the workbook
            try:
                wb.save(output_file)
                self.logger.info(f"Excel file saved successfully: {output_file}")
                return True
            except PermissionError:
                # Try with a different filename if file is locked
                alt_name = f"alt_{int(time.time())}_{os.path.basename(output_file)}"
                alternative_file = os.path.join(os.path.dirname(output_file), alt_name)
                self.logger.warning(f"Permission error saving Excel file, trying alternative: {alternative_file}")
                
                try:
                    wb.save(alternative_file)
                    self.logger.info(f"Excel file saved with alternative name: {alternative_file}")
                    return True
                except Exception as e2:
                    self.logger.error(f"Failed to save with alternative name: {str(e2)}")
                    import traceback
                    self.logger.error(traceback.format_exc())
                    return False
                
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            
            # Try a minimal version if full version fails
            try:
                self._save_minimal_excel(output_file)
                return True
            except Exception as e2:
                self.logger.error(f"Error saving minimal Excel file: {str(e2)}")
                return False

    def _save_minimal_excel(self, output_file):
        """Save a minimal version of the results to Excel in case the full version fails."""
        try:
            self.logger.info(f"Attempting to save minimal Excel file to {output_file}")
            
            # Don't rely on calculate_summary_statistics, use self.results directly
            # This avoids NoneType is not subscriptable errors
            
            # Create a list to hold material data
            materials_list = []
            
            # Iterate through materials in self.results if it exists
            if isinstance(self.results, dict) and 'materials' in self.results:
                for mat_name, mat_data in self.results['materials'].items():
                    # Get volume and count with safe defaults
                    volume = 0
                    count = 0
                    
                    # Safely get values using get() with defaults
                    if isinstance(mat_data, dict):
                        volume = mat_data.get('total_volume', 0)
                        count = mat_data.get('count', 0)
                    
                    # Add to our materials list
                    materials_list.append({
                        'Material': mat_name,
                        'Volume (m³)': volume,
                        'Count': count
                    })
            
            # If we have no materials, add a placeholder row
            if not materials_list:
                materials_list.append({
                    'Material': 'No materials found',
                    'Volume (m³)': 0,
                    'Count': 0
                })
            
            # Check if the directory exists
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                self.logger.info(f"Created directory: {output_dir}")
            
            # Check if directory is writable
            try:
                test_file = os.path.join(output_dir, '.test_write')
                with open(test_file, 'w') as f:
                    f.write('test')
                os.remove(test_file)
                self.logger.info(f"Directory {output_dir} is writable")
            except Exception as e:
                self.logger.error(f"Directory {output_dir} is not writable: {str(e)}")
                
                # Try to use a temp directory instead
                import tempfile
                temp_dir = tempfile.gettempdir()
                output_file = os.path.join(temp_dir, os.path.basename(output_file))
                self.logger.info(f"Using temp directory instead: {output_file}")
            
            # Create DataFrame and save to Excel
            import pandas as pd
            df = pd.DataFrame(materials_list)
            df.to_excel(output_file, sheet_name='Summary', index=False)
            
            # Verify the file was created
            if os.path.exists(output_file):
                self.logger.info(f"Minimal Excel file saved: {output_file}")
                return True
            else:
                self.logger.error(f"Failed to create Excel file at {output_file}")
                return False
            
        except Exception as e:
            self.logger.error(f"Failed to save minimal Excel: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            
            # Last resort: try to save as CSV instead
            try:
                csv_file = os.path.join(os.path.dirname(output_file), f"{os.path.splitext(os.path.basename(output_file))[0]}.csv")
                self.logger.info(f"Attempting to save as CSV instead: {csv_file}")
                
                # Create a simple DataFrame with error information
                import pandas as pd
                pd.DataFrame([{'Error': str(e)}]).to_csv(csv_file, index=False)
                
                # Verify the CSV file was created
                if os.path.exists(csv_file):
                    self.logger.info(f"Data saved as CSV instead: {csv_file}")
                    return True
                else:
                    self.logger.error(f"Failed to create CSV file at {csv_file}")
                    return False
            except Exception as csv_error:
                self.logger.error(f"Failed to save CSV file: {str(csv_error)}")
                return False

    def _create_element_sheet(self, wb):
        """Create the Element Type Summary sheet in the Excel workbook."""
        element_sheet = wb.create_sheet("Element Type Summary")
        
        # Add headers for Element Type Summary
        headers = [
            'Element Type', 'Count', 'Total Volume (m³)', 
            'Total Area (m²)', 'Total Weight (kg)', 'Materials'
        ]
        
        # Style the headers
        for col, header in enumerate(headers, 1):
            cell = element_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Add data to Element Type Summary sheet
        row = 2
        for element_type, data in self.results['element_types'].items():
            if data['count'] > 0:
                # Calculate total weight (using default density of steel)
                total_weight = data['total_volume'] * 7850  # Default to steel density
                
                # Write data to separate cells
                element_sheet.cell(row=row, column=1, value=element_type)
                element_sheet.cell(row=row, column=2, value=data['count'])
                element_sheet.cell(row=row, column=3, value=round(data['total_volume'], 3))
                element_sheet.cell(row=row, column=4, value=round(data['total_area'], 3))
                element_sheet.cell(row=row, column=5, value=round(total_weight, 1))
                element_sheet.cell(row=row, column=6, value=', '.join(data['materials'].keys()))
                
                # Add border to all cells in the row
                for col in range(1, len(headers) + 1):
                    element_sheet.cell(row=row, column=col).border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'),
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                
                row += 1
        
        # Adjust column widths
        self._adjust_column_widths(element_sheet)
        
    def _create_material_sheet(self, wb):
        """Create the Material Summary sheet in the Excel workbook."""
        summary_sheet = wb.create_sheet("Material Summary")
        
        # Add headers for Material Summary
        headers = [
            'Material Name', 'Total Count', 'Total Volume (m³)', 
            'Total Area (m²)', 'Total Weight (kg)', 'Grade', 
            'Specification', 'Material Type', 'Category', 'Description'
        ]
        
        # Style the headers
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Add data to Material Summary sheet
        row = 2
        for material_name, data in self.results['materials'].items():
            # Calculate density for weight calculation (default to 7850 kg/m³ for steel if not specified)
            density = 7850  # Default density (steel)
            for prop_name, prop_value in data['properties'].items():
                if 'density' in prop_name.lower() and prop_value:
                    try:
                        density = float(prop_value)
                        break
                    except ValueError:
                        pass
            
            # Calculate total weight
            total_weight = data['total_volume'] * density
            
            # Write data to separate cells
            summary_sheet.cell(row=row, column=1, value=material_name)
            summary_sheet.cell(row=row, column=2, value=data['count'])
            summary_sheet.cell(row=row, column=3, value=round(data['total_volume'], 3))
            summary_sheet.cell(row=row, column=4, value=round(data['total_area'], 3))
            summary_sheet.cell(row=row, column=5, value=round(total_weight, 1))
            summary_sheet.cell(row=row, column=6, value=', '.join(sorted(data['grades'])))
            summary_sheet.cell(row=row, column=7, value=', '.join(sorted(data['specifications'])))
            summary_sheet.cell(row=row, column=8, value=data['material_type'])
            summary_sheet.cell(row=row, column=9, value=data['category'])
            summary_sheet.cell(row=row, column=10, value=data['description'])
            
            # Add border to all cells in the row
            for col in range(1, len(headers) + 1):
                summary_sheet.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            row += 1
        
        # Adjust column widths
        self._adjust_column_widths(summary_sheet)
        
    def _create_summary_sheet(self, wb):
        """Create the detailed Material Takeoff sheet in the Excel workbook."""
        # Make the main takeoff sheet the active one
        if "Sheet" in wb.sheetnames:
            takeoff_sheet = wb["Sheet"]
            takeoff_sheet.title = "Detailed Material Takeoff"
        else:
            takeoff_sheet = wb.create_sheet("Detailed Material Takeoff", 0)
        
        # Add headers for Material Takeoff (professional format)
        headers = [
            'Element Type', 'Material Name', 'Grade', 'Specification', 
            'Length (m)', 'Width (m)', 'Height (m)', 'Quantity', 
            'Unit', 'Volume Each (m³)', 'Total Volume (m³)', 
            'Weight Each (kg)', 'Total Weight (kg)', 'Comments'
        ]
        
        # Style the headers
        for col, header in enumerate(headers, 1):
            cell = takeoff_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Check if we have element catalog data
        if 'element_catalog' not in self.results:
            takeoff_sheet.cell(row=2, column=1, value="No detailed element data available")
            return
            
        # Add data to Material Takeoff sheet
        row = 2
        current_element_type = None
        
        # First sort the catalog by element type, then by material
        try:
            sorted_keys = sorted(self.results['element_catalog'].keys(), 
                              key=lambda x: (x.split('|')[0], x.split('|')[1]))
        except (KeyError, ValueError):
            # Fallback if the key format is unexpected
            sorted_keys = list(self.results['element_catalog'].keys())
        
        for key in sorted_keys:
            try:
                element_type, material_name, _ = key.split('|', 2)
                data = self.results['element_catalog'][key]
                material_data = data['material_data']
                dimensions = data['dimensions']
                
                # Create element type headers (groups)
                if current_element_type != element_type:
                    current_element_type = element_type
                    takeoff_sheet.cell(row=row, column=1, value=element_type)
                    takeoff_sheet.cell(row=row, column=1).font = Font(bold=True)
                    for col in range(1, len(headers) + 1):
                        takeoff_sheet.cell(row=row, column=col).fill = PatternFill(
                            start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
                        )
                    row += 1
                
                # Calculate density for weight calculation (default to 7850 kg/m³ for steel if not specified)
                density = 7850  # Default density (steel)
                for prop_name, prop_value in material_data['properties'].items():
                    if 'density' in prop_name.lower() and prop_value:
                        try:
                            density = float(prop_value)
                            break
                        except ValueError:
                            pass
                
                # Calculate weight
                volume_each = data['volume'] / data['count'] if data['count'] > 0 else 0
                weight_each = volume_each * density
                total_weight = data['volume'] * density
                
                # Prepare comments (can include material type, category, etc.)
                comments = material_data['description']
                
                # Write data
                takeoff_sheet.cell(row=row, column=1, value=element_type)
                takeoff_sheet.cell(row=row, column=2, value=material_name)
                takeoff_sheet.cell(row=row, column=3, value=', '.join(material_data['grades']))
                takeoff_sheet.cell(row=row, column=4, value=', '.join(material_data['specifications']))
                takeoff_sheet.cell(row=row, column=5, value=dimensions['length'])
                takeoff_sheet.cell(row=row, column=6, value=dimensions['width'])
                takeoff_sheet.cell(row=row, column=7, value=dimensions['height'])
                takeoff_sheet.cell(row=row, column=8, value=data['count'])
                takeoff_sheet.cell(row=row, column=9, value="ea")
                takeoff_sheet.cell(row=row, column=10, value=round(volume_each, 3))
                takeoff_sheet.cell(row=row, column=11, value=round(data['volume'], 3))
                takeoff_sheet.cell(row=row, column=12, value=round(weight_each, 1))
                takeoff_sheet.cell(row=row, column=13, value=round(total_weight, 1))
                takeoff_sheet.cell(row=row, column=14, value=comments)
                
                # Add border to all cells in the row
                for col in range(1, len(headers) + 1):
                    takeoff_sheet.cell(row=row, column=col).border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'),
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                
                row += 1
            except Exception as e:
                self.logger.warning(f"Error processing row for {key}: {str(e)}")
                continue
        
        # Adjust column widths
        self._adjust_column_widths(takeoff_sheet)
        
    def _adjust_column_widths(self, worksheet):
        """Adjust column widths based on cell content."""
        try:
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                # Limit column width to avoid extremely wide columns
                adjusted_width = min(adjusted_width, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        except Exception as e:
            self.logger.warning(f"Error adjusting column widths: {str(e)}")
            # Don't raise the exception to avoid breaking the entire process

def main():
    """Main function to run the material takeoff analyzer."""
    if len(sys.argv) != 2:
        logger.error("Please provide an IFC file path")
        logger.info("Usage: python material_takeoff.py path/to/your_model.ifc")
        return 1
    
    ifc_file_path = sys.argv[1]
    if not os.path.isfile(ifc_file_path):
        logger.error(f"IFC file not found: {ifc_file_path}")
        return 1
    
    try:
        # Create analyzer and process elements
        analyzer = MaterialTakeoffAnalyzer(ifc_file_path)
        results = analyzer.analyze_all_elements()
        
        # Save results in all formats
        analyzer.save_results('all')
        
        # Display summary
        logger.info("\nMaterial Takeoff Summary:")
        for element_type, data in results['element_types'].items():
            if data['count'] > 0:
                logger.info(f"\n{element_type}:")
                logger.info(f"  Count: {data['count']}")
                logger.info(f"  Total Volume: {data['total_volume']:.2f} m³")
                logger.info(f"  Total Area: {data['total_area']:.2f} m²")
                logger.info("  Materials:")
                for material_name, material_data in data['materials'].items():
                    logger.info(f"    - {material_name}: {material_data['count']} elements")
                    if material_data['grades']:
                        logger.info(f"      Grades: {', '.join(material_data['grades'])}")
        
        return 0
    
    except Exception as e:
        logger.exception(f"Error analyzing IFC file: {e}")
        return 1

if __name__ == "__main__":
    sys.exit(main()) 