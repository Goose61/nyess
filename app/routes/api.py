import os
import json
import time
import traceback
import threading
from flask import (
    Blueprint, request, current_app, jsonify, abort, url_for, render_template
)
from werkzeug.utils import secure_filename
from app.models.material_takeoff import MaterialTakeoffAnalyzer
from app.routes.main import analysis_tasks, analyze_file_task, threads
import openpyxl
from openpyxl.utils import get_column_letter
import copy
import csv
from openpyxl import Workbook
from openpyxl.styles import Font

bp = Blueprint('api', __name__, url_prefix='/api')

def allowed_file(filename):
    """Check if the file has an allowed extension."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

@bp.route('/status/<path:filename>', methods=['GET'])
def get_analysis_status(filename):
    """Get the status of an analysis task."""
    try:
        # Sanitize filename to prevent path traversal
        filename = secure_filename(filename)
        
        # Log important information for debugging
        current_app.logger.info(f"Checking status for file: {filename}")
        current_app.logger.info(f"Upload folder: {current_app.config['UPLOAD_FOLDER']}")
        current_app.logger.info(f"Results folder: {current_app.config.get('RESULTS_FOLDER')}")
        
        # Check if the analysis task exists
        if filename not in analysis_tasks:
            # No task found - create a pending task instead of showing an error
            current_app.logger.info(f"Analysis task not found for {filename}, creating a new pending task")
            
            # Initialize a new task - this will be picked up by normal loading flow
            analysis_tasks[filename] = {
                'status': 'pending',
                'error': None,
                'phase': 'initializing',
                'phase_description': 'Preparing to analyze file',
                'start_time': time.time()
            }
            
            # Return a pending status instead of error
            return jsonify({
                'status': 'pending',
                'message': 'Analysis task initialized',
                'html': render_template('loading_status.html', 
                                      status='pending',
                                      task=analysis_tasks[filename],
                                      filename=filename)
            })
        
        # Get the task status
        task = analysis_tasks[filename]
        current_app.logger.info(f"Task status for {filename}: {task['status']}")
        
        if task['status'] == 'completed':
            # Set redirect URL
            task['redirect_url'] = url_for('main.analyze', filename=filename)
            
            return jsonify({
                'status': 'completed',
                'redirect_url': url_for('main.analyze', filename=filename),
                'total_elements': task.get('total_elements', 0),
                'processed_elements': task.get('total_elements', 0),
                'html': render_template('loading_status.html', 
                                      status='completed',
                                      task=task,
                                      filename=filename)
            })
        elif task['status'] == 'failed':
            # Analysis failed
            return jsonify({
                'status': 'failed',
                'error': task.get('error', 'Unknown error'),
                'phase': task.get('phase', 'error'),
                'phase_description': task.get('phase_description', 'Analysis failed'),
                'html': render_template('loading_status.html', 
                                      status='failed',
                                      task=task,
                                      filename=filename)
            })
        else:
            # Analysis in progress
            response_data = {
                'status': task['status'],
                'phase': task.get('phase', 'running'),
                'phase_description': task.get('phase_description', 'Analysis in progress')
            }
            
            # Include timing information if available
            if 'start_time' in task:
                elapsed_time = time.time() - task['start_time']
                response_data['elapsed_time'] = elapsed_time
                # Format elapsed time for display
                hours, remainder = divmod(elapsed_time, 3600)
                minutes, seconds = divmod(remainder, 60)
                task['elapsed_time_formatted'] = f"{int(hours)}h {int(minutes)}m {int(seconds)}s" if hours > 0 else f"{int(minutes)}m {int(seconds)}s"
            
            # Only include element count information if available
            if 'total_elements' in task and 'processed_elements' in task:
                total = task['total_elements']
                processed = task['processed_elements']
                
                # Check if we're in the detailed processing phase
                if 'detailed_processing_elements' in task:
                    # Use the detailed processing count when that data is available
                    detailed_total = task.get('total_elements', 0)
                    detailed_processed = task.get('detailed_processing_elements', 0)
                    
                    # Only override if we have actual data
                    if detailed_processed > 0:
                        processed = detailed_processed
                        
                        # Update phase info if not already set
                        if task.get('phase') == 'generating_results':
                            response_data['phase_description'] = f'Processing detailed element data: {detailed_processed}/{detailed_total} ({(detailed_processed/detailed_total)*100:.1f}%)'
                
                response_data.update({
                    'total_elements': total,
                    'processed_elements': processed
                })
                
                # Calculate processing rate and estimated time remaining
                if processed > 0 and elapsed_time > 0 and processed < total:
                    # Calculate processing rate (elements per second)
                    processing_rate = processed / elapsed_time
                    
                    # Estimate time remaining
                    if processing_rate > 0:
                        remaining_elements = total - processed
                        estimated_seconds_remaining = remaining_elements / processing_rate
                        
                        # Format for display
                        hours, remainder = divmod(estimated_seconds_remaining, 3600)
                        minutes, seconds = divmod(remainder, 60)
                        
                        # Add to task and response
                        task['processing_rate'] = processing_rate
                        task['processing_rate_formatted'] = f"{processing_rate:.1f} elements/sec"
                        task['estimated_time_remaining'] = f"{int(hours)}h {int(minutes)}m {int(seconds)}s" if hours > 0 else f"{int(minutes)}m {int(seconds)}s"
                        
                        response_data['processing_rate'] = processing_rate
                        response_data['estimated_time_remaining'] = task['estimated_time_remaining']
                
                # If we have element processing info and in the generating results phase
                if task.get('phase') == 'generating_results' and task.get('element_processing_time'):
                    response_data['element_processing_time'] = task.get('element_processing_time')
                    response_data['element_processing_complete'] = True
            
            # Add HTML content to the response for AJAX updates
            response_data['html'] = render_template('loading_status.html', 
                                                  status=task['status'],
                                                  task=task,
                                                  filename=filename,
                                                  time=time)
            
            return jsonify(response_data)
    
    except Exception as e:
        # Log the error but don't expose details to client
        current_app.logger.error(f"Error retrieving status for {filename}: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': 'Internal server error',
            'html': render_template('loading_status.html', 
                                  status='failed',
                                  task={'error': 'Internal server error'},
                                  filename=filename)
        }), 500

@bp.route('/upload', methods=['POST'])
def upload_file():
    """API endpoint for file upload."""
    try:
        # Check if the post request has the file part
        if 'file' not in request.files:
            return jsonify({'error': 'No file part'}), 400
        
        file = request.files['file']
        
        # If user does not select file, browser also
        # submit an empty part without filename
        if file.filename == '':
            return jsonify({'error': 'No selected file'}), 400
        
        if file and allowed_file(file.filename):
            # Secure the filename and add timestamp to avoid overwrites
            timestamp = int(time.time())
            filename = f"{timestamp}_{secure_filename(file.filename)}"
            file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            
            # Ensure upload directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # Save the file
            file.save(file_path)
            
            current_app.logger.info(f"File uploaded: {filename}")
            
            # Initialize analysis task status
            analysis_tasks[filename] = {
                'status': 'pending',
                'error': None,
                'results': None
            }
            
            return jsonify({
                'success': True,
                'message': 'File uploaded successfully',
                'filename': filename,
                'loading_url': url_for('main.loading', filename=filename)
            })
        
        else:
            return jsonify({
                'error': 'File type not allowed. Please upload an IFC file.'
            }), 400
    except Exception as e:
        current_app.logger.error(f"Error uploading file: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': 'Internal server error during file upload'}), 500

@bp.route('/analyze/<filename>', methods=['GET'])
def analyze(filename):
    """API endpoint to analyze an uploaded IFC file."""
    try:
        # Sanitize filename to prevent path traversal
        filename = os.path.basename(filename)
        
        # Check if file exists and hasn't been analyzed yet
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(file_path):
            current_app.logger.warning(f"File not found: {filename}")
            return jsonify({'error': 'File not found'}), 404
        
        # Start or check analysis task
        if filename not in analysis_tasks:
            # Initialize task
            analysis_tasks[filename] = {
                'status': 'pending',
                'error': None,
                'results': None
            }
            
            # Get upload folder and app for the background thread
            upload_folder = current_app.config['UPLOAD_FOLDER']
            app_instance = current_app._get_current_object()
            
            # Start analysis in background thread
            analysis_tasks[filename]['status'] = 'running'
            
            # Check if thread already exists
            if filename in threads and threads[filename].is_alive():
                current_app.logger.warning(f"Thread for {filename} already exists. Status update only.")
            else:
                # Create and start new thread
                thread = threading.Thread(
                    target=analyze_file_task,
                    args=(filename, upload_folder, app_instance),
                    name=f"analysis_{filename}"
                )
                thread.daemon = True
                thread.start()
                threads[filename] = thread
                current_app.logger.info(f"Started analysis thread from API for {filename}")
            
            # Redirect to loading endpoint
            return jsonify({
                'status': 'pending',
                'message': 'Analysis task created',
                'loading_url': url_for('main.loading', filename=filename)
            })
        
        # Get task status
        task = analysis_tasks.get(filename, {})
        status = task.get('status', 'unknown')
        
        if status == 'completed':
            # Return results
            return jsonify({
                'success': True,
                'message': 'Analysis completed',
                'files': task.get('results', {})
            })
        elif status == 'failed':
            # Return error
            error_msg = task.get('error', 'Unknown error')
            return jsonify({
                'error': f'Analysis failed: {error_msg}'
            }), 500
        else:
            # Analysis in progress
            return jsonify({
                'status': status,
                'message': 'Analysis in progress',
                'total_elements': task.get('total_elements', 0),
                'processed_elements': task.get('processed_elements', 0)
            })
    
    except Exception as e:
        current_app.logger.error(f"Error analyzing file: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': 'Internal server error during analysis'}), 500

@bp.route('/results/<filename>', methods=['GET'])
def get_results(filename):
    """API endpoint to get analysis results."""
    try:
        # Sanitize filename to prevent path traversal
        filename = os.path.basename(filename)
        
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        
        # Check if file exists
        if not os.path.exists(file_path):
            current_app.logger.warning(f"Results file not found: {filename}")
            return jsonify({'error': 'File not found'}), 404
        
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)
        except (IOError, PermissionError) as e:
            current_app.logger.error(f"Error opening file {filename}: {str(e)}")
            return jsonify({'error': f'Error opening file: {str(e)}'}), 500
        
        return jsonify(data)
    
    except json.JSONDecodeError as e:
        current_app.logger.error(f"Invalid JSON in file: {filename} - {str(e)}")
        return jsonify({'error': f'Invalid JSON format: {str(e)}'}), 400
        
    except Exception as e:
        current_app.logger.error(f"Error reading results: {filename} - {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': f'Error reading results: {str(e)}'}), 500

@bp.route('/generate_excel/<filename>', methods=['GET'])
def generate_excel(filename):
    """Generate Excel file with adjusted quantities."""
    try:
        # Sanitize filename to prevent path traversal
        filename = os.path.basename(filename)
        
        # Get the adjustment percentage from query parameters
        adjustment = float(request.args.get('adjustment', 100)) / 100
        
        # Check if file exists
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        # Generate a unique filename for the Excel file
        timestamp = int(time.time())
        excel_filename = f"{timestamp}_{os.path.splitext(filename)[0]}_material_takeoff.xlsx"
        excel_path = os.path.join(current_app.config['UPLOAD_FOLDER'], excel_filename)
        
        # Read the JSON data
        json_filename = f"{os.path.splitext(filename)[0]}_analysis.json"
        json_path = os.path.join(current_app.config['UPLOAD_FOLDER'], json_filename)
        
        with open(json_path, 'r') as f:
            data = json.load(f)
        
        # Create Excel workbook
        wb = Workbook()
        
        # Create summary sheet - ensure it exists
        summary_sheet = wb.active
        if summary_sheet is None:
            summary_sheet = wb.create_sheet("Summary")
        else:
            summary_sheet.title = "Summary"
        
        # Add headers
        headers = ["Element Type", "Material", "Count", "Volume (m³)", "Area (m²)"]
        for col, header in enumerate(headers, 1):
            summary_sheet.cell(row=1, column=col, value=header)
        
        # Add data with adjusted quantities
        row = 2
        for element_type, materials in data.items():
            for material, props in materials.items():
                summary_sheet.cell(row=row, column=1, value=element_type)
                summary_sheet.cell(row=row, column=2, value=material)
                summary_sheet.cell(row=row, column=3, value=props['count'])
                summary_sheet.cell(row=row, column=4, value=props['volume'] * adjustment)
                summary_sheet.cell(row=row, column=5, value=props['area'] * adjustment)
                row += 1
        
        # Format the sheet
        for col in range(1, 6):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in summary_sheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(excel_path)
        
        # Return the download URL
        return jsonify({
            'success': True,
            'download_url': url_for('main.download_file', filename=excel_filename),
            'filename': excel_filename
        })
        
    except Exception as e:
        current_app.logger.error(f"Error generating Excel: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': f'Error generating Excel file: {str(e)}'}), 500 

@bp.route('/export/<filename>', methods=['GET'])
def export_file(filename):
    """Export analysis results in various formats with quantity adjustments."""
    try:
        # Get export parameters
        format_type = request.args.get('format', 'json')
        adjustment = float(request.args.get('adjustment', 100)) / 100  # Convert percentage to multiplier
        
        # Validate adjustment value (prevent negative or unreasonable values)
        if adjustment <= 0 or adjustment > 10:  # Cap at 1000%
            return jsonify({'error': 'Invalid adjustment value. Must be between 0 and 1000%'}), 400
        
        # Sanitize filename
        filename = secure_filename(filename)
        
        # Check if the analysis task exists
        if filename not in analysis_tasks:
            return jsonify({'error': 'Analysis task not found'}), 404
        
        # Get task information
        task = analysis_tasks[filename]
        
        if task['status'] != 'completed':
            return jsonify({'error': 'Analysis not completed'}), 400
        
        # Get result files information
        results = task.get('results', {})
        
        # Get JSON file path (required for all exports)
        json_file = results.get('json_file', '')
        json_path = os.path.join(current_app.config['UPLOAD_FOLDER'], json_file)
        
        if not os.path.exists(json_path):
            return jsonify({'error': 'JSON result file missing'}), 404
        
        # Load the JSON data
        try:
            with open(json_path, 'r') as f:
                original_data = json.load(f)
        except Exception as e:
            current_app.logger.error(f"Error reading JSON file: {str(e)}")
            return jsonify({'error': f'Error reading JSON data: {str(e)}'}), 500
        
        # Create a new data structure with adjusted quantities
        adjusted_data = adjust_quantities(original_data, adjustment)
        
        # Process based on the requested format
        if format_type == 'json':
            # Generate adjusted JSON file
            adjusted_json_file = f"{os.path.splitext(json_file)[0]}_adjusted.json"
            adjusted_json_path = os.path.join(current_app.config['UPLOAD_FOLDER'], adjusted_json_file)
            
            try:
                with open(adjusted_json_path, 'w') as f:
                    json.dump(adjusted_data, f, indent=4)
            except Exception as e:
                current_app.logger.error(f"Error writing adjusted JSON file: {str(e)}")
                return jsonify({'error': f'Error creating adjusted JSON file: {str(e)}'}), 500
            
            return jsonify({
                'success': True,
                'filename': adjusted_json_file,
                'download_url': url_for('main.download_file', filename=adjusted_json_file)
            })
            
        elif format_type == 'excel':
            # Generate adjusted Excel file
            base_name = os.path.splitext(filename)[0]
            excel_file = f"{base_name}_material_takeoff_adjusted.xlsx"
            excel_path = os.path.join(current_app.config['UPLOAD_FOLDER'], excel_file)
            
            try:
                generate_excel_from_data(adjusted_data, excel_path)
            except Exception as e:
                current_app.logger.error(f"Error generating Excel file: {str(e)}")
                return jsonify({'error': f'Error generating Excel file: {str(e)}'}), 500
            
            return jsonify({
                'success': True,
                'filename': excel_file,
                'download_url': url_for('main.download_file', filename=excel_file)
            })
            
        elif format_type == 'csv':
            # Generate adjusted CSV files
            base_name = os.path.splitext(filename)[0]
            summary_file = f"{base_name}_material_takeoff_summary_adjusted.csv"
            details_file = f"{base_name}_material_takeoff_details_adjusted.csv"
            
            summary_path = os.path.join(current_app.config['UPLOAD_FOLDER'], summary_file)
            details_path = os.path.join(current_app.config['UPLOAD_FOLDER'], details_file)
            
            try:
                generate_csv_from_data(adjusted_data, summary_path, details_path)
            except Exception as e:
                current_app.logger.error(f"Error generating CSV files: {str(e)}")
                return jsonify({'error': f'Error generating CSV files: {str(e)}'}), 500
            
            return jsonify({
                'success': True,
                'filename': summary_file,
                'download_url': url_for('main.download_file', filename=summary_file)
            })
            
        elif format_type == 'all':
            # Generate all formats
            base_name = os.path.splitext(filename)[0]
            
            # Create adjusted filenames
            adjusted_json_file = f"{base_name}_material_takeoff_adjusted.json"
            adjusted_excel_file = f"{base_name}_material_takeoff_adjusted.xlsx"
            adjusted_summary_file = f"{base_name}_material_takeoff_summary_adjusted.csv"
            adjusted_details_file = f"{base_name}_material_takeoff_details_adjusted.csv"
            
            # Create file paths
            adjusted_json_path = os.path.join(current_app.config['UPLOAD_FOLDER'], adjusted_json_file)
            adjusted_excel_path = os.path.join(current_app.config['UPLOAD_FOLDER'], adjusted_excel_file)
            adjusted_summary_path = os.path.join(current_app.config['UPLOAD_FOLDER'], adjusted_summary_file)
            adjusted_details_path = os.path.join(current_app.config['UPLOAD_FOLDER'], adjusted_details_file)
            
            files_to_generate = []
            
            # Add JSON to the files list
            try:
                with open(adjusted_json_path, 'w') as f:
                    json.dump(adjusted_data, f, indent=4)
                files_to_generate.append({
                    'filename': adjusted_json_file,
                    'url': url_for('main.download_file', filename=adjusted_json_file)
                })
            except Exception as e:
                current_app.logger.error(f"Error generating adjusted JSON file: {str(e)}")
            
            # Add Excel to the files list
            try:
                generate_excel_from_data(adjusted_data, adjusted_excel_path)
                files_to_generate.append({
                    'filename': adjusted_excel_file,
                    'url': url_for('main.download_file', filename=adjusted_excel_file)
                })
            except Exception as e:
                current_app.logger.error(f"Error generating adjusted Excel file: {str(e)}")
            
            # Add CSV to the files list
            try:
                generate_csv_from_data(adjusted_data, adjusted_summary_path, adjusted_details_path)
                files_to_generate.append({
                    'filename': adjusted_summary_file,
                    'url': url_for('main.download_file', filename=adjusted_summary_file)
                })
                files_to_generate.append({
                    'filename': adjusted_details_file,
                    'url': url_for('main.download_file', filename=adjusted_details_file)
                })
            except Exception as e:
                current_app.logger.error(f"Error generating adjusted CSV files: {str(e)}")
            
            if not files_to_generate:
                return jsonify({'error': 'Failed to generate any adjusted files'}), 500
            
            return jsonify({
                'success': True,
                'files': files_to_generate
            })
            
        else:
            return jsonify({'error': 'Invalid format type'}), 400
        
    except ValueError as e:
        return jsonify({'error': f'Invalid adjustment value: {str(e)}'}), 400
    except Exception as e:
        current_app.logger.error(f"Error in export route: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

def adjust_quantities(data, adjustment_factor):
    """
    Adjust quantities in the data by the given factor.
    
    Args:
        data (dict): Original data from JSON
        adjustment_factor (float): Factor to adjust quantities by (e.g., 1.1 for 10% increase)
        
    Returns:
        dict: Adjusted data
    """
    adjusted_data = copy.deepcopy(data)
    
    # Adjust material volumes and areas
    for material_name, material_data in adjusted_data.get('materials', {}).items():
        material_data['total_volume'] *= adjustment_factor
        material_data['total_area'] *= adjustment_factor
    
    # Adjust element type volumes and areas
    for element_type, element_data in adjusted_data.get('element_types', {}).items():
        element_data['total_volume'] *= adjustment_factor
        element_data['total_area'] *= adjustment_factor
        
        # Adjust materials within element types
        for material_name, material_data in element_data.get('materials', {}).items():
            material_data['volume'] *= adjustment_factor
            material_data['area'] *= adjustment_factor
    
    # Adjust element catalog if present
    if 'element_catalog' in adjusted_data:
        for key, catalog_item in adjusted_data['element_catalog'].items():
            catalog_item['volume'] *= adjustment_factor
            catalog_item['area'] *= adjustment_factor
    
    return adjusted_data

def generate_excel_from_data(data, output_path):
    """Generate Excel file from adjusted data."""
    wb = Workbook()
    
    # Create Material Summary sheet - ensure it exists
    summary_sheet = wb.active
    if summary_sheet is None:
        summary_sheet = wb.create_sheet("Material Summary")
    else:
        summary_sheet.title = "Material Summary"
    
    # Add headers
    headers = ['Material Name', 'Count', 'Total Volume (m³)', 'Total Area (m²)', 
              'Total Weight (kg)', 'Category', 'Material Type', 'Description']
    
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add material data
    row = 2
    for material_name, material_data in data.get('materials', {}).items():
        if material_data['count'] > 0:
            # Calculate weight (using default density for steel if not specified)
            density = 7850
            for prop_name, prop_value in material_data.get('properties', {}).items():
                if 'density' in prop_name.lower() and prop_value:
                    try:
                        density = float(prop_value)
                        break
                    except ValueError:
                        pass
            
            weight = material_data['total_volume'] * density
            
            summary_sheet.cell(row=row, column=1, value=material_name)
            summary_sheet.cell(row=row, column=2, value=material_data['count'])
            summary_sheet.cell(row=row, column=3, value=round(material_data['total_volume'], 3))
            summary_sheet.cell(row=row, column=4, value=round(material_data['total_area'], 3))
            summary_sheet.cell(row=row, column=5, value=round(weight, 1))
            summary_sheet.cell(row=row, column=6, value=material_data.get('category', ''))
            summary_sheet.cell(row=row, column=7, value=material_data.get('material_type', ''))
            summary_sheet.cell(row=row, column=8, value=material_data.get('description', ''))
            
            row += 1
    
    # Create Element Types sheet
    element_sheet = wb.create_sheet("Element Types")
    
    # Add headers
    headers = ['Element Type', 'Count', 'Total Volume (m³)', 'Total Area (m²)', 'Materials']
    
    for col, header in enumerate(headers, 1):
        cell = element_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add element type data
    row = 2
    for element_type, element_data in data.get('element_types', {}).items():
        if element_data['count'] > 0:
            element_sheet.cell(row=row, column=1, value=element_type)
            element_sheet.cell(row=row, column=2, value=element_data['count'])
            element_sheet.cell(row=row, column=3, value=round(element_data['total_volume'], 3))
            element_sheet.cell(row=row, column=4, value=round(element_data['total_area'], 3))
            element_sheet.cell(row=row, column=5, value=', '.join(element_data.get('materials', {}).keys()))
            
            row += 1
    
    # Adjust column widths for all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_cells:  # Ensure column is not empty
                adjusted_width = (max_length + 2)
                column_letter = column_cells[0].column_letter
                ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_path)

def generate_csv_from_data(data, summary_path, details_path):
    """Generate CSV files from adjusted data."""
    # Generate summary CSV
    with open(summary_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Material Name', 'Count', 'Total Volume (m³)', 
                        'Total Area (m²)', 'Total Weight (kg)', 'Category'])
        
        for material_name, material_data in data.get('materials', {}).items():
            if material_data['count'] > 0:
                # Calculate weight
                density = 7850
                for prop_name, prop_value in material_data.get('properties', {}).items():
                    if 'density' in prop_name.lower() and prop_value:
                        try:
                            density = float(prop_value)
                            break
                        except ValueError:
                            pass
                
                weight = material_data['total_volume'] * density
                
                writer.writerow([
                    material_name,
                    material_data['count'],
                    f"{material_data['total_volume']:.3f}",
                    f"{material_data['total_area']:.3f}",
                    f"{weight:.1f}",
                    material_data.get('category', '')
                ])
    
    # Generate details CSV
    with open(details_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Element Type', 'Count', 'Volume (m³)', 
                        'Area (m²)', 'Materials'])
        
        for element_type, element_data in data.get('element_types', {}).items():
            if element_data['count'] > 0:
                writer.writerow([
                    element_type,
                    element_data['count'],
                    f"{element_data['total_volume']:.3f}",
                    f"{element_data['total_area']:.3f}",
                    ', '.join(element_data.get('materials', {}).keys())
                ]) 